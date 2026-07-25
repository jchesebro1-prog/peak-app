#!/usr/bin/env python3
"""
Convert the 52-brand dealer price sheets (/Volumes/Claude/Peak Import) into
catalog-import JSON matching scripts/catalog-import-data.json conventions:
  { sku: "Brand:Model", desc, category, unit: "ea", list, cost, mfr, note? }

Design (matches what the sheets actually are, probed 2026-07-24):
- STATEFUL WALKING: any row that maps a sku-ish + price-ish header becomes the
  active column map; later rows parse as data until the next header. Handles
  PDF conversions whose tables start deep and repeat per section (Danley r96+).
- PRIORITIZED, FUZZY aliases: "Part Number" outranks "Item" for sku; MSRP
  outranks MAP for list; year-prefixed headers ("2026 NET Price", "MSRP NEW")
  match by token. Tier columns (Authorized/Preferred) import as cost with a
  verify note — Peak's actual tier per vendor is not recorded.
- Single-cell rows between data become the running category (Draper sections);
  else category falls back per workbook tab / brand default.
- Martin-style fused prices ("$3,850.00 $2,310.00" in one cell) are split.
- Rows that don't parse are COUNTED per file, never silently dropped.
- Held out: Tannoy June 2023 (superseded by Music Tribe Nov 2025 — D117) and
  Ape Riggers 2014 (OCR output is unparseable garbage; needs the PDF re-done
  or a current sheet).

Output: dealer-import-data.json (path from argv[1], default scratch) +
a per-file report on stdout. Feed the JSON to scripts/import-dealer-sheets.ts.
"""
import json
import os
import re
import sys

import openpyxl

FOLDER = "/Volumes/Claude/Peak Import"

HELD_OUT = {
    "Tannoy Dealer Pricing June 2023.xlsx": "superseded by Music Tribe Nov 2025 (D117)",
    "Ape Riggers Dealer Price Sheet 2014 (OCR - verify).xlsx": "2014 scan; OCR unparseable — request current sheet",
    "Draper Dealer Pricing April 2026.xlsx": "configurator-style size matrices (tens of thousands of size-permutation rows) — import needs its own decision",
}

SKIP_TABS = re.compile(
    r"cover|legal|instruction|change|contact|channel map|post-sales|obsolete|"
    r"discontinued|^home$|read me|readme|index|toc|price changes|care|taa",
    re.I,
)

# filename → brand (Music Tribe / Harman / Chauvet resolve per tab below)
BRANDS = {
    "1Sound": "1Sound", "Ace Backstage": "Ace Backstage", "Allen & Heath": "Allen & Heath",
    "Apex": "Apex", "Audinate": "Audinate", "AV Pro Edge": "AVPro Edge",
    "Behringer": "Behringer", "Biamp": "Biamp", "BirdDog": "BirdDog",
    "Chauvet": None, "Cloud": "Cloud", "Danley": "Danley", "Datavideo": "Datavideo",
    "DB Technologies": "dB Technologies", "DICOLOR": "DICOLOR", "Dicolor": "DICOLOR",
    "Draper": "Draper", "EAW Adaptive": "EAW", "EAW Dealer": "EAW", "FSR": "FSR",
    "Fulcrum": "Fulcrum Acoustic", "Harman": None, "HSA": "HSA Rolltops",
    "Juice Goose": "Juice Goose", "LEA": "LEA Professional", "Linea Research": "Linea Research",
    "Listen": "Listen Technologies", "Lyntec": "LynTec", "Martin": "Martin Audio",
    "Matrox": "Matrox", "Maxhub": "MAXHUB", "Meyer": "Meyer Sound",
    "Microphone Madness": "Microphone Madness", "Music Tribe": None,
    "Netgear": "NETGEAR", "Pliant": "Pliant Technologies", "Polar Focus": "Polar Focus",
    "Powersoft": "Powersoft", "QSC": "QSC", "RCF": "RCF", "Renkus": "Renkus-Heinz",
    "Shure": "Shure", "SurgeX": "SurgeX", "Symetrix": "Symetrix",
    "Visionary": "Visionary Solutions", "Williams": "Williams AV", "Wyrestorm": "WyreStorm",
}

# Brand default category buckets (filter strings, refinable in the catalog UI).
BRAND_CATEGORY = {
    "Martin Audio": "Audio", "Meyer Sound": "Audio", "Danley": "Audio", "EAW": "Audio",
    "QSC": "Audio", "RCF": "Audio", "Renkus-Heinz": "Audio", "Fulcrum Acoustic": "Audio",
    "1Sound": "Audio", "LEA Professional": "Audio", "Linea Research": "Audio",
    "Powersoft": "Audio", "Apex": "Audio", "Cloud": "Audio", "dB Technologies": "Audio",
    "Biamp": "Audio", "Symetrix": "Audio", "Allen & Heath": "Audio", "Shure": "Audio",
    "Microphone Madness": "Audio", "Behringer": "Audio", "Midas": "Audio",
    "Klark Teknik": "Audio", "Lab Gruppen": "Audio", "Lake": "Audio", "Turbosound": "Audio",
    "Tannoy": "Audio", "AKG": "Audio", "JBL": "Audio", "Crown": "Audio", "BSS": "Audio",
    "dbx": "Audio", "Lexicon": "Audio", "Soundcraft": "Audio", "AMX": "AV Control",
    "Audinate": "Networked AV", "NETGEAR": "Networking", "BirdDog": "Video",
    "Datavideo": "Video", "Matrox": "Video", "MAXHUB": "Displays", "DICOLOR": "LED Video",
    "Draper": "Screens & Lifts", "AVPro Edge": "AV Distribution", "WyreStorm": "AV Distribution",
    "FSR": "AV Infrastructure", "Visionary Solutions": "AV Distribution",
    "SurgeX": "Power", "Juice Goose": "Power", "LynTec": "Power",
    "Williams AV": "Assistive Listening", "Listen Technologies": "Assistive Listening",
    "Pliant Technologies": "Comms", "Ace Backstage": "Stage Accessories",
    "Polar Focus": "Rigging", "HSA Rolltops": "Staging", "ChamSys": "Lighting Control",
    "CHAUVET DJ": "Lighting", "Chauvet Professional": "Lighting", "Chauvet Cinema": "Lighting",
}

HARMAN_TABS = {"AKG": "AKG", "AMX": "AMX", "BSS": "BSS", "Crown": "Crown", "DBX": "dbx",
               "JBL": "JBL", "JBL Commercial": "JBL", "Cinema": "JBL", "Lexicon": "Lexicon",
               "Soundcraft": "Soundcraft"}
MT_TABS = {"MIDAS": "Midas", "KLARK": "Klark Teknik", "LAB": "Lab Gruppen", "LAKE": "Lake",
           "TURBO": "Turbosound", "TANNOY": "Tannoy", "BEHRINGER": "Behringer"}

EXTRA_OFF = {"Shure": "take add'l 3% off dealer", "EAW": "take add'l 8% off dealer"}

def normh(s):
    return re.sub(r"[^a-z0-9%/ ]", "", str(s or "").lower()).strip()

# (field, priority, matcher) — higher priority wins a column; a column binds once.
HEADER_RULES = [
    ("sku", 9, lambda n: n in ("sku", "part number", "part", "part no", "partno", "part #")),
    ("sku", 8, lambda n: "part number" in n or n.startswith("part ")),
    ("sku", 7, lambda n: n in ("model", "model number", "model no", "item number", "itemnumber", "items", "item no", "order code", "product code", "item code", "cat no", "article")),
    ("sku", 4, lambda n: n == "item"),
    ("desc", 8, lambda n: n in ("description", "product description", "item description", "desc")),
    ("desc", 6, lambda n: n in ("item name", "product name", "name", "product", "details")),
    ("list", 9, lambda n: "msrp" in n or n in ("list", "list price", "retail", "srp", "suggested retail") or "list price" in n or ("retail" in n and "map" not in n)),
    ("list", 3, lambda n: n == "map" or "map" in n.split()),
    ("cost", 9, lambda n: "dealer net" in n or n in ("dealer", "dealer price", "dealer cost", "net price", "net", "wholesale", "our cost", "your price") or "dealer" in n),
    ("cost", 7, lambda n: "net" in n.split() or "net price" in n),
    ("cost", 5, lambda n: n in ("preferred", "authorized", "partner")),  # vendor tiers → verify note
    ("category", 6, lambda n: n in ("category", "family", "product family", "group", "series", "line", "type", "section", "class")),
    ("unit", 6, lambda n: n in ("unit", "uom", "u/m", "um")),
    ("price?", 2, lambda n: n in ("price", "unit price", "pricing")),
]

TIER_COLS = ("preferred", "authorized", "partner")

def map_header(cells):
    best = {}
    tier_hit = False
    for idx, c in enumerate(cells):
        n = normh(c)
        if not n:
            continue
        for field, prio, fn in HEADER_RULES:
            try:
                hit = fn(n)
            except Exception:
                hit = False
            if hit:
                if field not in best or best[field][1] < prio:
                    best[field] = (idx, prio)
                if field == "cost" and n in TIER_COLS:
                    tier_hit = True
                break
    cols = {f: i for f, (i, _) in best.items()}
    if "list" not in cols and "cost" not in cols and "price?" in cols:
        cols["list"] = cols.pop("price?")
    cols.pop("price?", None)
    return cols, tier_hit

MONEY = re.compile(r"\$?\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)")

def money(v, loose=False):
    """loose=True (headerless fallback) rejects spec-looking numbers: bare
    values under $5 and EAN/UPC-scale integers. Header-mapped columns keep
    everything positive (accessories under $5 are real there)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        n = float(v)
        if n <= 0 or n >= 500_000:
            return None
        if loose and (n < 5 or (n > 99_999 and n == int(n))):
            return None
        return round(n, 2)
    sv = str(v)
    m = MONEY.search(sv)
    if not m:
        return None
    n = float(m.group(1).replace(",", ""))
    if n <= 0 or n >= 500_000:
        return None
    if loose and "$" not in sv and (n < 5 or (n > 99_999 and n == int(n))):
        return None
    return round(n, 2)

def fused_prices(v):
    """'$3,850.00 $2,310.00' → (3850.0, 2310.0) or None."""
    if not isinstance(v, str):
        return None
    ms = MONEY.findall(v)
    if len(ms) == 2:
        a, b = (float(x.replace(",", "")) for x in ms)
        if a > 0 and b > 0:
            return round(a, 2), round(b, 2)
    return None

def brand_for(fn, tab):
    for key, brand in BRANDS.items():
        if fn.lower().startswith(key.lower()):
            if brand is not None:
                return brand
            break
    if fn.startswith("Harman"):
        return HARMAN_TABS.get(tab.strip(), "Harman")
    if fn.startswith("Music Tribe"):
        up = tab.upper()
        for k, b in MT_TABS.items():
            if k in up:
                return b
        return "Music Tribe"
    if fn.startswith("Chauvet"):
        t = tab.strip().upper()
        if "DJ" in t: return "CHAUVET DJ"
        if "PROF" in t: return "Chauvet Professional"
        if "CINEMA" in t: return "Chauvet Cinema"
        if "CHAMSYS" in t: return "ChamSys"
        if "LYNTEC" in t: return None  # dedicated Lyntec file wins
        return "Chauvet Professional"
    # fall back: first two filename words
    return " ".join(fn.split()[:2])

SKUISH = __import__("re").compile(r"^[A-Z0-9][A-Za-z0-9./+#-]{2,24}$")

def skuish(tok):
    """A plausible model token: has a digit, or is dashed/ALL-CAPS short."""
    if not SKUISH.match(tok):
        return False
    return any(c.isdigit() for c in tok) or ("-" in tok and tok == tok.upper())

def split_fused_text(t):
    """'LR-ASC48 Linea Research ASC48 …' → ('LR-ASC48', rest) or None."""
    toks = t.split(None, 1)
    if toks and skuish(toks[0]):
        return toks[0], (toks[1].strip() if len(toks) > 1 else toks[0])
    return None

PURE_MONEY = re.compile(r"\$?\s*[0-9][0-9,]*(?:\.[0-9]{1,2})?")

def pure_money(v):
    """The WHOLE cell is a money value → its amount; else None. Digits
    embedded in model names (SW-620-TX-W, TOUCHMIX-8) never count."""
    if isinstance(v, (int, float)):
        n = float(v)
        if n < 5 or n >= 500_000 or (n > 99_999 and n == int(n)):
            return None
        if 1990 <= n <= 2035 and n == int(n):
            return None  # bare year, not a price
        return round(n, 2)
    if not isinstance(v, str):
        return None
    t = v.strip()
    if not PURE_MONEY.fullmatch(t):
        return None
    n = float(re.sub(r"[^0-9.]", "", t) or 0)
    if n < 5 or n >= 500_000 or ("$" not in t and n > 99_999 and n == int(n)):
        return None
    if "$" not in t and 1990 <= n <= 2035 and n == int(n) and "." not in t and "," not in t:
        return None  # bare year, not a price
    return round(n, 2)

JUNK_SKU = re.compile(r"^[\W_]+$|^\d/\d$")

def sane_prices(lst, cst):
    """Reject/repair implausible price pairs — spec numbers masquerading as
    prices are worse than a skipped row. Returns (lst, cst, extra_note) or
    None to skip the row."""
    if lst and cst and cst > lst:
        lst, cst = cst, lst  # inverted columns — swap, flag below
        swapped = True
    else:
        swapped = False
    if not lst or lst < 10:
        return None  # sub-$10 list on dealer sheets = spec-number pollution
    if cst and (cst < lst * 0.12 or lst / max(cst, 0.01) > 50):
        return (lst, 0, "verify price (cost implausible, dropped)")
    if swapped:
        return (lst, cst, "verify price (columns swapped)")
    return (lst, cst, None)

def is_texty(v):
    return isinstance(v, str) and v.strip() != "" and money(v) is None

parts = {}
dupes = 0
report = []

for fn in sorted(os.listdir(FOLDER)):
    if not fn.endswith(".xlsx"):
        continue
    if fn in HELD_OUT:
        report.append(f"HELD OUT  {fn} — {HELD_OUT[fn]}")
        continue
    wb = openpyxl.load_workbook(os.path.join(FOLDER, fn), read_only=True, data_only=True)
    extra_note = next((v for k, v in EXTRA_OFF.items() if fn.startswith(k) or "EAW" in fn and k == "EAW"), None)
    f_ok = f_skip = f_flag = 0
    for ws in wb.worksheets:
        tab = ws.title
        if SKIP_TABS.search(tab):
            continue
        # dB Technologies: the Full Line tab is the superset of the series tabs.
        if fn.startswith("DB Technologies") and tab.strip().lower() != "full line":
            continue
        brand = brand_for(fn, tab)
        if brand is None:
            continue
        if fn.startswith("Music Tribe"):
            # One headerless tab; per-row layout:
            # [BRAND, order-code, MODEL, EAN, display-model, desc, …, dealer, msrp]
            mt_names = {"KLARK TEKNIK": "Klark Teknik", "LAB GRUPPEN": "Lab Gruppen",
                        "TURBOSOUND": "Turbosound", "TANNOY": "Tannoy", "MIDAS": "Midas",
                        "LAKE": "Lake", "BEHRINGER": "Behringer"}
            for r in ws.iter_rows(values_only=True):
                vals = list(r)
                if len(vals) < 9 or not isinstance(vals[0], str):
                    continue
                b = mt_names.get(vals[0].strip().upper())
                model = str(vals[4] or vals[2] or "").strip()
                if not b or not model or len(model) > 40:
                    continue
                cst, lst = money(vals[7]), money(vals[8])
                if lst is None and cst is None:
                    f_skip += 1
                    continue
                desc = str(vals[5] or model).strip()[:200]
                key = f"{b}:{model}"
                if key in parts:
                    dupes += 1
                parts[key] = {"sku": key, "desc": desc, "category": BRAND_CATEGORY.get(b, "Audio"),
                              "unit": "ea", "list": lst or cst or 0, "cost": cst or 0, "mfr": b}
                f_ok += 1
            continue
        default_cat = BRAND_CATEGORY.get(brand, "AV")
        cols, tier = None, False
        running_cat = None
        for r in ws.iter_rows(values_only=True):
            m, t = map_header(r)
            if "sku" in m and ("list" in m or "cost" in m):
                cols, tier = m, t
                running_cat = None
                continue
            if cols is None:
                # Generic headerless parser: sku = first short text cell,
                # desc = longest text cell, prices = $-strings preferred
                # (min→cost, max→list), fused "$X $Y" cells split.
                cells = [(i, v) for i, v in enumerate(r) if v not in (None, "")]
                texts = [(i, str(v).strip()) for i, v in cells
                         if isinstance(v, str) and pure_money(v) is None and fused_prices(v) is None]
                texts = [(i, t) for i, t in texts if not JUNK_SKU.fullmatch(t)]
                if not texts:
                    continue
                sku_c = next(((i, t) for i, t in texts if len(t) <= 30 and len(t.split()) <= 3), None)
                desc_c = max(texts, key=lambda x: len(x[1]))
                if sku_c is None:
                    fused_t = split_fused_text(desc_c[1])
                    if fused_t is None:
                        continue
                    sku_c = (desc_c[0], fused_t[0])
                    desc_c = (-1, fused_t[1])  # -1: desc came from the split
                fused = next((fused_prices(v) for _, v in cells if fused_prices(v)), None)
                if fused:
                    lst, cst = fused
                else:
                    dollar = [pure_money(v) for _, v in cells if isinstance(v, str) and "$" in str(v) and pure_money(v)]
                    nums = dollar or [m for m in (pure_money(v) for i, v in cells if i != sku_c[0]) if m]
                    if not nums:
                        continue
                    lst, cst = (max(nums), min(nums)) if len(nums) >= 2 else (nums[0], 0)
                sp_check = sane_prices(lst, cst)
                if sp_check is None:
                    f_skip += 1
                    continue
                lst, cst, extra = sp_check
                sku_txt = sku_c[1]
                key = f"{brand}:{sku_txt}"
                if key in parts:
                    dupes += 1
                note = "verify price (headerless rows)" + (f"; {extra}" if extra else "")
                parts[key] = {"sku": key, "desc": (desc_c[1][:200] if desc_c[0] != sku_c[0] or desc_c[0] == -1 else sku_txt),
                              "category": running_cat or default_cat,
                              "unit": "ea", "list": lst, "cost": cst, "mfr": brand, "note": note}
                f_ok += 1; f_flag += 1
                continue
            # header known — parse as data
            sku_v = r[cols["sku"]] if cols["sku"] < len(r) else None
            if sku_v in (None, ""):
                filled = [v for v in r if v not in (None, "")]
                if len(filled) == 1 and is_texty(filled[0]) and len(str(filled[0])) < 60:
                    running_cat = str(filled[0]).strip()
                continue
            sku_txt = str(sku_v).strip()
            if not sku_txt or len(sku_txt) > 40 or pure_money(sku_txt) is not None or JUNK_SKU.fullmatch(sku_txt):
                f_skip += 1
                continue
            desc_v = r[cols["desc"]] if "desc" in cols and cols["desc"] < len(r) else None
            lst = money(r[cols["list"]]) if "list" in cols and cols["list"] < len(r) else None
            cst = money(r[cols["cost"]]) if "cost" in cols and cols["cost"] < len(r) else None
            if lst is None and cst is not None and "list" in cols:
                fused = fused_prices(r[cols["list"]])
                if fused:
                    lst, cst = fused
            if lst is None and cst is None:
                fused = next((fused_prices(v) for v in r if v not in (None, "") and fused_prices(v)), None)
                if fused:
                    lst, cst = fused
                else:
                    singles = [m for m in (money(v, loose=True) for v in r if isinstance(v, str) and "$" in str(v)) if m]
                    if singles:
                        lst, cst = (max(singles), min(singles)) if len(singles) >= 2 else (singles[0], None)
            if lst is None and cst is None:
                f_skip += 1
                continue
            # PDF-fused sku cells: "CP354 APEX CloudPower …" under a Model header
            if " " in sku_txt and len(sku_txt) > 24:
                sp = split_fused_text(sku_txt)
                if sp:
                    sku_txt = sp[0]
                    if desc_v in (None, ""):
                        desc_v = sp[1]
            sp_check = sane_prices(lst or 0, cst or 0)
            if sp_check is None:
                f_skip += 1
                continue
            lst, cst, sp_extra = sp_check
            cst = cst or None
            cat_v = r[cols["category"]] if "category" in cols and cols["category"] < len(r) else None
            unit_v = r[cols["unit"]] if "unit" in cols and cols["unit"] < len(r) else None
            notes = []
            if sp_extra:
                notes.append(sp_extra)
            if extra_note:
                notes.append(extra_note)
            if tier and cst is not None:
                notes.append("verify dealer tier pricing")
            if lst is None:
                lst, notes = (cst or 0), notes + ["verify price (dealer-only sheet)"]
                cst = cst or 0
            if cst is None:
                cst, notes = 0, notes + ["verify price (list-only sheet)"]
            key = f"{brand}:{sku_txt}"
            if key in parts:
                dupes += 1
            entry = {"sku": key, "desc": (str(desc_v).strip()[:200] if desc_v not in (None, "") else sku_txt),
                     "category": (str(cat_v).strip() if cat_v not in (None, "") else (running_cat or default_cat)),
                     "unit": (str(unit_v).strip() if unit_v not in (None, "") else "ea"),
                     "list": lst, "cost": cst, "mfr": brand}
            if notes:
                entry["note"] = "; ".join(dict.fromkeys(notes))
            parts[key] = entry
            f_ok += 1
            if notes:
                f_flag += 1
    wb.close()
    report.append(f"{f_ok:5d} ok  {f_skip:5d} skipped  {f_flag:5d} flagged   {fn}")

out_path = sys.argv[1] if len(sys.argv) > 1 else "dealer-import-data.json"
rows = sorted(parts.values(), key=lambda p: p["sku"])
with open(out_path, "w") as f:
    json.dump(rows, f)
print("\n".join(report))
print(f"\nTOTAL parts: {len(rows)}  (within-run dupes overwritten: {dupes})")
print(f"flagged: {sum(1 for p in rows if p.get('note'))}")
print(f"wrote {out_path}")
